"""Reconcile FIL + ETH trades in DB against the latest Excel workbook.

Dry-run: identifies trades that exist in the Excel but are missing from the DB.
Does NOT modify the database.
"""
import sqlite3
from datetime import datetime, date
from collections import defaultdict
import openpyxl

TODAY = date.today().isoformat()
EXCEL_PATH = r"C:\Users\Lucas Lemos\Downloads\Copy of FIL, ETH & BTC Option Strategies (2).xlsx"
DB_PATH = r"C:\Users\Lucas Lemos\PycharmProjects\plgo_options\data\plgo_options.db"

# Sheet -> (asset, counterparty-as-stored-in-DB)
SHEETS = {
    "GSR - FIL Option Positions (New":    ("FIL", "GSR"),
    "Wave - FIL Option Positions (Ne":    ("FIL", "Wave"),
    "G20 - FIL Option Positions":         ("FIL", "G20"),
    "Flowdesk - FIL Option Positions":    ("FIL", "Flowdesk"),
    "Keyrock - FIL Option Positions":     ("FIL", "KeyRock"),
    "GSR - ETH Option Positions":         ("ETH", "GSR"),
    "Wave - ETH Option Positions":        ("ETH", "Wave"),
    "Flowdesk - ETH Option Positions":    ("ETH", "FlowDesk"),
    "Keyrock - ETH Option Positions":     ("ETH", "KeyRock"),
}


def to_date_str(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return str(v).split("T")[0] if v else ""


def to_float(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except (ValueError, TypeError):
        return 0.0


def norm_side(s):
    return str(s or "").strip().lower()


def norm_opt(s):
    return str(s or "").strip().lower().rstrip("s")  # calls -> call, puts -> put


def make_key(asset, cpty, trade_date, side, option_type, expiry, strike, qty):
    return (
        asset,
        cpty,
        trade_date.split("T")[0],
        norm_side(side),
        norm_opt(option_type),
        expiry.split("T")[0],
        round(float(strike), 4),
        round(float(qty), 0),
    )


# ── Load DB trades ──
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
db_trades = [dict(r) for r in conn.execute(
    "SELECT * FROM trades WHERE status != 'deleted' ORDER BY id"
).fetchall()]
conn.close()

db_keys = defaultdict(list)
for t in db_trades:
    k = make_key(t["asset"], t["counterparty"], t["trade_date"], t["side"],
                 t["option_type"], t["expiry"], t["strike"], t["qty"])
    db_keys[k].append(t)

print(f"DB total (non-deleted): {len(db_trades)}")

# ── Parse Excel ──
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
excel_trades = []

for sname, (asset, cpty) in SHEETS.items():
    if sname not in wb.sheetnames:
        print(f"WARNING: sheet missing: {sname}")
        continue
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, r in enumerate(rows):
        if r and r[0] and str(r[0]).strip() == "Initial Trade Date":
            header_idx = i
            break
    if header_idx is None:
        print(f"WARNING: No header in {sname}")
        continue

    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(rows[header_idx])]

    # Qty column name: FIL uses "# of FIL Options", ETH uses "# of ETH Options"
    qty_col = f"# of {asset} Options"
    # Strike column: "$ Strike" for most, but might vary
    strike_col = "$ Strike"
    # Premium columns
    prem_total_col = "Total $ Premium Received/(Paid)"
    prem_per_col = "Premium/ option ($)"

    for i in range(header_idx + 1, len(rows)):
        r = rows[i]
        if not r or all(c is None for c in r):
            continue
        rec = dict(zip(headers, r))

        td = rec.get("Initial Trade Date")
        side = rec.get("Buy / Sell / Unwind")
        opt_type = rec.get("Option Type")
        expiry = rec.get("Option Expiry Date")
        strike = rec.get(strike_col)
        qty = rec.get(qty_col)

        if not td or not side or not opt_type:
            continue
        side_s = norm_side(side)
        if side_s in ("", "total", "totals"):
            continue

        strike_val = to_float(strike)
        qty_val = to_float(qty)
        if strike_val <= 0 or qty_val <= 0:
            continue

        excel_trades.append({
            "asset": asset,
            "counterparty": cpty,
            "trade_date": to_date_str(td),
            "side": str(side).strip(),
            "option_type": str(opt_type).strip(),
            "expiry": to_date_str(expiry),
            "strike": strike_val,
            "qty": qty_val,
            "ref_spot": to_float(rec.get("Ref. Spot Price")),
            "pct_otm": to_float(rec.get("% OTM")),
            "notional_mm": to_float(rec.get("$ Notional (mm)")),
            "premium_per": to_float(rec.get(prem_per_col)),
            "premium_usd": to_float(rec.get(prem_total_col)),
            "_sheet": sname,
        })
wb.close()

print(f"Excel parsed: {len(excel_trades)} trades across {len(SHEETS)} sheets")

# ── Match: multiset by key ──
xl_by_key = defaultdict(list)
for t in excel_trades:
    k = make_key(t["asset"], t["counterparty"], t["trade_date"], t["side"],
                 t["option_type"], t["expiry"], t["strike"], t["qty"])
    xl_by_key[k].append(t)

missing = []
for k, xl_list in xl_by_key.items():
    db_count = len(db_keys.get(k, []))
    for t in xl_list[db_count:]:
        missing.append(t)

# ── Classify missing ──
missing.sort(key=lambda t: (t["asset"], t["counterparty"], t["trade_date"], t["expiry"]))

by_asset = defaultdict(list)
for t in missing:
    by_asset[t["asset"]].append(t)

print(f"\n=== MISSING FROM DB: {len(missing)} total ===")
for asset in ("FIL", "ETH"):
    rows = by_asset.get(asset, [])
    active = [t for t in rows if t["expiry"] >= TODAY]
    expired = [t for t in rows if t["expiry"] < TODAY]
    print(f"  {asset}: {len(rows)} total ({len(active)} active, {len(expired)} expired)")

# Focus on recent trades (last 3 weeks) — what the user cares about
THRESHOLD = "2026-04-01"
recent = [t for t in missing if t["trade_date"] >= THRESHOLD]
print(f"\n=== RECENT MISSING (trade_date >= {THRESHOLD}): {len(recent)} ===")
for t in sorted(recent, key=lambda x: (x["asset"], x["trade_date"], x["counterparty"], x["expiry"])):
    status = "active" if t["expiry"] >= TODAY else "expired"
    print(f"  {t['asset']:3s} | {t['counterparty']:8s} | td={t['trade_date']} | "
          f"{t['side']:6s} {t['option_type']:5s} | K={t['strike']:<10} | "
          f"exp={t['expiry']} | qty={t['qty']:<12} | prem=${t['premium_usd']:<12} | {status}")

# Also show older missing, bucketed
older = [t for t in missing if t["trade_date"] < THRESHOLD]
if older:
    print(f"\n=== OLDER MISSING (trade_date < {THRESHOLD}): {len(older)} ===")
    by_ac = defaultdict(int)
    for t in older:
        by_ac[(t["asset"], t["counterparty"])] += 1
    for (a, c), n in sorted(by_ac.items()):
        print(f"  {a} / {c}: {n}")
