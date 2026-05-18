"""Insert the 56 recent missing FIL + ETH trades (trade_date >= 2026-04-01)
from the latest Excel workbook into plgo_options.db.

Writes to the trades table with status='active' and logs audit entries
tagged 'excel_import_recon_20260420'. Matches existing per-asset
conventions: FIL stores side lowercase / option_type plural lowercase,
ETH stores side Capitalized / option_type Singular capitalized.
"""
import sqlite3
from datetime import datetime, date
from collections import defaultdict
import openpyxl

CUTOFF = "2026-04-01"
TODAY = date.today().isoformat()
AUDIT_TAG = "excel_import_recon_20260420"
EXCEL_PATH = r"C:\Users\Lucas Lemos\Downloads\Copy of FIL, ETH & BTC Option Strategies (2).xlsx"
DB_PATH = r"C:\Users\Lucas Lemos\PycharmProjects\plgo_options\data\plgo_options.prod.db"

SHEETS = {
    "GSR - FIL Option Positions (New":    ("FIL", "GSR"),
    "Wave - FIL Option Positions (Ne":    ("FIL", "Wave"),
    "G20 - FIL Option Positions":         ("FIL", "G20"),
    "Flowdesk - FIL Option Positions":    ("FIL", "Flowdesk"),
    "Keyrock - FIL Option Positions":     ("FIL", "KeyRock"),
    "GSR - ETH Option Positions":         ("ETH", "GSR"),
    "Wave - ETH Option Positions":        ("ETH", "Wave"),
    "Flowdesk - ETH Option Positions":    ("ETH", "FlowDesk"),
    "Keyrock - ETH Option Positions":     ("ETH", "KeyRock"),
}


def to_date_str(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return str(v).split("T")[0] if v else ""


def to_float(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except (ValueError, TypeError):
        return 0.0


def norm_side(s):
    return str(s or "").strip().lower()


def norm_opt(s):
    return str(s or "").strip().lower().rstrip("s")


def match_key(asset, cpty, td_, side, ot, exp, strike, qty):
    return (asset, cpty, td_.split("T")[0], norm_side(side), norm_opt(ot),
            exp.split("T")[0], round(float(strike), 4), round(float(qty), 0))


# Formats that match existing asset-specific conventions in DB
def fmt_side(asset, side):
    s = norm_side(side)
    if asset == "ETH":
        return s.capitalize()
    return s  # FIL stays lowercase


def fmt_opt(asset, ot):
    s = str(ot or "").strip().lower()
    if asset == "ETH":
        return s.rstrip("s").capitalize()  # Calls -> Call
    if not s.endswith("s"):
        s = s + "s"
    return s  # FIL stays plural lowercase


# ── Load existing DB trades & build multiset of keys ──
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
db_rows = conn.execute(
    "SELECT id, asset, counterparty, trade_date, side, option_type, expiry, strike, qty "
    "FROM trades WHERE status != 'deleted'"
).fetchall()
db_key_count = defaultdict(int)
for r in db_rows:
    k = match_key(r["asset"], r["counterparty"], r["trade_date"], r["side"],
                  r["option_type"], r["expiry"], r["strike"], r["qty"])
    db_key_count[k] += 1

# ── Parse Excel ──
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
excel_trades = []
for sname, (asset, cpty) in SHEETS.items():
    if sname not in wb.sheetnames:
        continue
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    hi = None
    for i, r in enumerate(rows):
        if r and r[0] and str(r[0]).strip() == "Initial Trade Date":
            hi = i
            break
    if hi is None:
        continue
    headers = [str(h).strip() if h else f"c{j}" for j, h in enumerate(rows[hi])]
    qty_col = f"# of {asset} Options"
    for i in range(hi + 1, len(rows)):
        r = rows[i]
        if not r or all(c is None for c in r):
            continue
        rec = dict(zip(headers, r))
        td_val = rec.get("Initial Trade Date")
        side = rec.get("Buy / Sell / Unwind")
        ot = rec.get("Option Type")
        if not td_val or not side or not ot:
            continue
        if norm_side(side) in ("", "total", "totals"):
            continue
        strike = to_float(rec.get("$ Strike"))
        qty = to_float(rec.get(qty_col))
        if strike <= 0 or qty <= 0:
            continue
        excel_trades.append({
            "asset": asset,
            "counterparty": cpty,
            "trade_date": to_date_str(td_val),
            "side_raw": side,
            "option_type_raw": ot,
            "expiry": to_date_str(rec.get("Option Expiry Date")),
            "strike": strike,
            "qty": qty,
            "ref_spot": to_float(rec.get("Ref. Spot Price")),
            "pct_otm": to_float(rec.get("% OTM")),
            "notional_mm": to_float(rec.get("$ Notional (mm)")),
            "premium_per": to_float(rec.get("Premium/ option ($)")),
            "premium_usd": to_float(rec.get("Total $ Premium Received/(Paid)")),
        })
wb.close()

# ── Identify missing trades in scope (trade_date >= CUTOFF) ──
xl_by_key = defaultdict(list)
for t in excel_trades:
    k = match_key(t["asset"], t["counterparty"], t["trade_date"], t["side_raw"],
                  t["option_type_raw"], t["expiry"], t["strike"], t["qty"])
    xl_by_key[k].append(t)

to_insert = []
for k, xl_list in xl_by_key.items():
    db_cnt = db_key_count.get(k, 0)
    for t in xl_list[db_cnt:]:
        if t["trade_date"] >= CUTOFF:
            to_insert.append(t)

to_insert.sort(key=lambda t: (t["asset"], t["trade_date"], t["counterparty"], t["expiry"]))

print(f"Will insert: {len(to_insert)} trades (cutoff trade_date >= {CUTOFF})")
by_ac = defaultdict(int)
for t in to_insert:
    by_ac[(t["asset"], t["counterparty"])] += 1
for (a, c), n in sorted(by_ac.items()):
    print(f"  {a} / {c}: {n}")

# ── Insert ──
inserted = 0
for t in to_insert:
    asset = t["asset"]
    status = "active" if t["expiry"] >= TODAY else "expired"
    side_db = fmt_side(asset, t["side_raw"])
    opt_db = fmt_opt(asset, t["option_type_raw"])

    cur = conn.execute(
        """INSERT INTO trades
           (asset, counterparty, trade_id, trade_date, side, option_type,
            expiry, strike, ref_spot, pct_otm, qty,
            notional_mm, premium_per, premium_usd, status)
           VALUES (?, ?, '', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (asset, t["counterparty"], t["trade_date"], side_db, opt_db,
         t["expiry"], t["strike"], t["ref_spot"], t["pct_otm"], t["qty"],
         t["notional_mm"], t["premium_per"], t["premium_usd"], status),
    )
    new_id = cur.lastrowid
    conn.execute(
        "INSERT INTO trade_audit_log (trade_id, action, changed_by) VALUES (?, 'create', ?)",
        (new_id, AUDIT_TAG),
    )
    inserted += 1

conn.commit()

# ── Verify ──
for asset in ("FIL", "ETH"):
    r = conn.execute(
        "SELECT COUNT(*) FROM trades WHERE asset=? AND status!='deleted'", (asset,)
    ).fetchone()[0]
    print(f"DB {asset} total now: {r}")

conn.close()
print(f"\n=== DONE: inserted {inserted} trades, audit tag '{AUDIT_TAG}' ===")
