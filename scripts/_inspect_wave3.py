import openpyxl
from collections import Counter
path = r'C:/Users/Lucas Lemos/Downloads/Filecoin - Wave.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
print("SHEETS:", wb.sheetnames)
for ws in wb.worksheets:
    print("\n\n######### SHEET:", ws.title, "max_row", ws.max_row, "max_col", ws.max_column)
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    # print first 12 rows
    for i, row in enumerate(rows[:12], 1):
        vals = ["" if v is None else str(v) for v in row]
        while vals and vals[-1] == "":
            vals.pop()
        print(f" R{i}: " + " | ".join(vals))
    # find header row + days-remaining col, bucket its values
    hdr = -1
    for i, row in enumerate(rows[:15]):
        if any(str(c or "").strip().lower() == "counterparty" for c in row):
            hdr = i; break
    if hdr < 0:
        print("  (no Counterparty header found)")
        continue
    header = [str(c or "").strip() for c in rows[hdr]]
    di = next((j for j, h in enumerate(header) if "days remaining" in h.lower()), None)
    print(f"  header row idx={hdr}; days-remaining col idx={di}")
    if di is not None:
        cnt = Counter()
        numeric_rows = 0
        for row in rows[hdr+1:]:
            v = row[di] if di < len(row) else None
            if v is None: cnt['(blank)'] += 1
            elif isinstance(v, (int, float)): cnt['NUMERIC'] += 1; numeric_rows += 1
            else: cnt[str(v)] += 1
        print("  days-remaining buckets:", cnt.most_common(12))
        print("  => OPEN (numeric days remaining):", numeric_rows)
