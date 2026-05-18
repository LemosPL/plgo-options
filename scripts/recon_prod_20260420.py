"""Reconcile the prod DB (pulled from GCS) against the latest Excel.

Only inspects. Prints a breakdown so we can decide what to insert.
"""
import sqlite3
from datetime import datetime, date
from collections import defaultdict
import openpyxl

TODAY = date.today().isoformat()
EXCEL_PATH = r"C:\Users\Lucas Lemos\Downloads\Copy of FIL, ETH & BTC Option Strategies (2).xlsx"
DB_PATH = r"C:\Users\Lucas Lemos\PycharmProjects\plgo_options\data\plgo_options.prod.db"

SHEETS = {
    "GSR - FIL Option Positions (New":    ("FIL", "GSR"),
    "Wave - FIL Option Positions (Ne":    ("FIL", "Wave"),
    "G20 - FIL Option Positions":         ("FIL", "G20"),
    "Flowdesk - FIL Option Positions":    ("FIL", "Flowdesk"),
    "Keyrock - FIL Option Positions":     ("FIL", "KeyRock"),
    "GSR - ETH Option Positions":         ("ETH", "GSR"),
    "Wave - ETH Option Positions":        ("ETH", "Wave"),
    "Flowdesk - ETH Option Positions":    ("ETH", "FlowDesk"),
    "Keyrock - ETH Option Positions":     ("ETH", "KeyRock"),
}


def td(v):
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    if isinstance(v, date): return v.isoformat()
    return str(v).split("T")[0] if v else ""
def tf(v):
    try: return float(v) if v not in (None,"") else 0.0
    except: return 0.0
def norm_side(s): return str(s or "").strip().lower()
def norm_opt(s): return str(s or "").strip().lower().rstrip("s")
def key(asset, cpty, t, s, o, e, k, q):
    return (asset, cpty, t.split("T")[0], norm_side(s), norm_opt(o), e.split("T")[0], round(float(k),4), round(float(q),0))


conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
db = [dict(r) for r in conn.execute("SELECT * FROM trades WHERE status!='deleted'")]
conn.close()
db_keys = defaultdict(int)
for t in db:
    db_keys[key(t['asset'], t['counterparty'], t['trade_date'], t['side'], t['option_type'], t['expiry'], t['strike'], t['qty'])] += 1

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
xl = []
for sname, (asset, cpty) in SHEETS.items():
    if sname not in wb.sheetnames: continue
    ws = wb[sname]; rows = list(ws.iter_rows(values_only=True))
    hi = None
    for i,r in enumerate(rows):
        if r and r[0] and str(r[0]).strip() == "Initial Trade Date":
            hi = i; break
    if hi is None: continue
    headers = [str(h).strip() if h else f"c{j}" for j,h in enumerate(rows[hi])]
    qty_col = f"# of {asset} Options"
    for i in range(hi+1, len(rows)):
        r = rows[i]
        if not r or all(c is None for c in r): continue
        rec = dict(zip(headers, r))
        tdate = rec.get("Initial Trade Date"); side = rec.get("Buy / Sell / Unwind"); ot = rec.get("Option Type")
        if not tdate or not side or not ot: continue
        if norm_side(side) in ("","total","totals"): continue
        strike = tf(rec.get("$ Strike")); qty = tf(rec.get(qty_col))
        if strike<=0 or qty<=0: continue
        xl.append({
            "asset": asset, "counterparty": cpty, "trade_date": td(tdate),
            "side_raw": side, "option_type_raw": ot,
            "expiry": td(rec.get("Option Expiry Date")), "strike": strike, "qty": qty,
            "ref_spot": tf(rec.get("Ref. Spot Price")), "pct_otm": tf(rec.get("% OTM")),
            "notional_mm": tf(rec.get("$ Notional (mm)")),
            "premium_per": tf(rec.get("Premium/ option ($)")),
            "premium_usd": tf(rec.get("Total $ Premium Received/(Paid)")),
        })
wb.close()

xl_keys = defaultdict(list)
for t in xl:
    xl_keys[key(t['asset'], t['counterparty'], t['trade_date'], t['side_raw'], t['option_type_raw'], t['expiry'], t['strike'], t['qty'])].append(t)

missing = []
for k, L in xl_keys.items():
    used = db_keys.get(k, 0)
    for t in L[used:]: missing.append(t)

print(f"Prod DB: {len(db)} | Excel: {len(xl)} | Missing: {len(missing)}")

# Break down missing by trade_date buckets
by_bucket = defaultdict(lambda: defaultdict(int))
for t in missing:
    td_ = t["trade_date"]
    if td_ >= "2026-04-01":
        bucket = "recent (>=2026-04-01)"
    elif td_ >= "2026-03-01":
        bucket = "march 2026"
    elif td_ >= "2026-01-01":
        bucket = "early 2026"
    else:
        bucket = "pre-2026"
    by_bucket[bucket][(t['asset'], t['counterparty'])] += 1

for bucket in ("recent (>=2026-04-01)", "march 2026", "early 2026", "pre-2026"):
    rows = by_bucket.get(bucket, {})
    total = sum(rows.values())
    if total == 0: continue
    print(f"\n  {bucket}: {total}")
    for (a,c), n in sorted(rows.items()):
        print(f"    {a} / {c}: {n}")

# Show the bucket of trades in 2026-03-07 to 2026-04-06 (the gap we expect)
print("\n=== TRADES IN GAP 2026-03-07..2026-04-06 (likely needed to catch prod up to local state): ===")
gap = [t for t in missing if "2026-03-07" <= t["trade_date"] <= "2026-04-06"]
print(f"Count: {len(gap)}")
for t in sorted(gap, key=lambda x: (x["trade_date"], x["asset"], x["counterparty"], x["expiry"])):
    status = "active" if t["expiry"] >= TODAY else "expired"
    print(f"  {t['asset']:3s} | {t['counterparty']:8s} | td={t['trade_date']} | "
          f"{t['side_raw']:6s} {t['option_type_raw']:5s} | K={t['strike']:<10} | "
          f"exp={t['expiry']} | qty={t['qty']:<12} | {status}")
