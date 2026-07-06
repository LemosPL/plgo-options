"""Read trades from the ETH/FIL dashboard Excel files."""

from __future__ import annotations

from pathlib import Path
from datetime import datetime, date

import openpyxl

# Expected columns from the 'Trades' sheet
TRADE_COLUMNS = [
    "Counterparty",
    "ID",
    "Initial Trade Date",
    "Buy / Sell / Unwind",
    "Option Type",
    "Trade_ID",
    "Option Expiry Date",
    "Days Remaining to Expiry",
    "Strike",
    "Ref. Spot Price",
    "% OTM",
    "ETH Options",
    "$ Notional (mm)",
    "Premium per Contract",
    "Premium USD",
]

def _latest_eth_xlsx() -> Path:
    """Return the most recently dated PLGO_Trades_*.xlsx in data/positions/."""
    positions_dir = Path(__file__).resolve().parents[3] / "data/positions"
    candidates = sorted(positions_dir.glob("PLGO_Trades_*.xlsx"))
    return candidates[-1] if candidates else positions_dir / "PLGO_Trades.xlsx"


# --- ETH paths ---
# Primary: latest PLGO_Trades_*.xlsx found in data/positions/
_PROJECT_DATA_PATH = _latest_eth_xlsx()

# Fallback: user's Downloads folder (local dev)
_DOWNLOADS_PATH = (
    Path.home() / "Downloads" / "PLGO_Trades_2026-05-26.xlsx"
)

# --- FIL paths ---
_FIL_PROJECT_DATA_PATH = (
    Path(__file__).resolve().parents[3]
    / "data"
    / "FIL - Dashboard Risk+PnL Improvement Proposal.xlsx"
)

_FIL_POSITIONS_DIR = (
    Path(__file__).resolve().parents[3]
    / "data"
    / "FIL_positions"
)

_FIL_DOWNLOADS_PATH = (
    Path.home() / "Downloads" / "FIL - Dashboard Risk+PnL Improvement Proposal (2).xlsx"
)


def _safe_float(v) -> float:
    """Coerce a value to float, returning 0.0 on failure."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# Maps the current 'Trades' sheet header names (ID, Status, Counterparty, Trade
# Date, Side, Type, Instrument, Expiry, DTE, Strike, % OTM, Qty, Notional ($mm),
# Premium USD, ...) onto the older canonical TRADE_COLUMNS names the rest of the
# app (portfolio.py, aggregate_positions, etc.) keys off of. Canonical names not
# present in this mapping (or already matching) pass through unchanged, so a
# sheet already using the old TRADE_COLUMNS headers still works.
_HEADER_ALIASES = {
    "Trade Date": "Initial Trade Date",
    "Side": "Buy / Sell / Unwind",
    "Type": "Option Type",
    "Instrument": "Trade_ID",
    "Expiry": "Option Expiry Date",
    "DTE": "Days Remaining to Expiry",
    "Qty": "ETH Options",
    "Notional ($mm)": "$ Notional (mm)",
}


def _normalize_trade_record(record: dict) -> dict:
    """Rename current-schema headers to the canonical TRADE_COLUMNS keys.

    Also derives 'Premium per Contract' and 'Ref. Spot Price' when the sheet
    doesn't carry them directly (the current export only has 'Premium USD').
    """
    normalized = {_HEADER_ALIASES.get(k, k): v for k, v in record.items()}

    if "Premium per Contract" not in normalized:
        qty = _safe_float(normalized.get("ETH Options"))
        premium_usd = _safe_float(normalized.get("Premium USD"))
        normalized["Premium per Contract"] = premium_usd / qty if qty else 0.0

    normalized.setdefault("Ref. Spot Price", 0.0)
    normalized.setdefault("$ Notional (mm)", 0.0)

    return normalized


def _read_trades_sheet(fp: Path, missing_header_msg: str) -> list[dict]:
    """Shared 'Trades' sheet parser for read_eth_trades / read_fil_trades."""
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws = wb["Trades"]
    rows = ws.iter_rows(values_only=True)

    # Scan for the header row containing "Counterparty", even if not in column A.
    headers: list[str] | None = None
    header_start_idx = 0
    for row in rows:
        if not row:
            continue

        normalized_cells = [
            str(cell).strip().lower() if cell is not None else ""
            for cell in row
        ]

        if "counterparty" in normalized_cells:
            header_start_idx = normalized_cells.index("counterparty")
            headers = [
                str(h).strip() if h is not None else f"col_{i}"
                for i, h in enumerate(row[header_start_idx:], start=header_start_idx)
            ]
            break

    if headers is None:
        wb.close()
        raise ValueError(missing_header_msg)

    trades: list[dict] = []
    for row in rows:
        if all(cell is None for cell in row):
            continue
        # row is NOT pre-sliced like headers is, so slice it the same way here —
        # otherwise fields silently bind to the wrong column when Counterparty
        # isn't in column A.
        record = dict(zip(headers, row[header_start_idx:]))
        # Normalise dates/datetimes → ISO strings
        for k, v in record.items():
            if isinstance(v, datetime):
                record[k] = v.isoformat()
            elif isinstance(v, date):
                record[k] = v.isoformat()
        trades.append(_normalize_trade_record(record))

    wb.close()
    return trades


def read_eth_trades(file_path: Path | None = None) -> list[dict]:
    """Return a list of trade dicts from the 'Trades' tab.

    Each dict is keyed by the canonical TRADE_COLUMNS names, regardless of
    whether the sheet uses the old or current header layout. Dates are
    normalised to ISO-8601 strings for JSON serialisation.
    """
    fp = file_path
    if fp is None:
        # Try bundled data/ first, then Downloads
        if _PROJECT_DATA_PATH.exists():
            fp = _PROJECT_DATA_PATH
        else:
            fp = _DOWNLOADS_PATH

    try:
        return _read_trades_sheet(
            fp,
            "Could not find header row starting with 'Counterparty' "
            "in the 'Trades' sheet",
        )
    except (FileNotFoundError, PermissionError, OSError):
        # Last resort: try the other path
        alt = _PROJECT_DATA_PATH if fp != _PROJECT_DATA_PATH else _DOWNLOADS_PATH
        if alt.exists():
            return _read_trades_sheet(
                alt,
                "Could not find header row starting with 'Counterparty' "
                "in the 'Trades' sheet",
            )
        raise


def _latest_trade_file(directory: Path) -> Path | None:
    files = sorted(directory.glob("PLGO_Trades_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def read_fil_trades(file_path: Path | None = None) -> list[dict]:
    """Return a list of trade dicts from the FIL 'Trades' tab.

    Same structure as read_eth_trades but reads from the FIL spreadsheet.
    The FIL spreadsheet reuses 'ETH Options' as the qty column header.
    """
    fp = file_path
    if fp is None:
        latest_positions_file = _latest_trade_file(_FIL_POSITIONS_DIR)
        if latest_positions_file is not None:
            fp = latest_positions_file
        elif _FIL_PROJECT_DATA_PATH.exists():
            fp = _FIL_PROJECT_DATA_PATH
        else:
            fp = _FIL_DOWNLOADS_PATH

    try:
        return _read_trades_sheet(
            fp,
            "Could not find header row starting with 'Counterparty' "
            "in the FIL 'Trades' sheet",
        )
    except (FileNotFoundError, PermissionError, OSError):
        alt = _FIL_PROJECT_DATA_PATH if fp != _FIL_PROJECT_DATA_PATH else _FIL_DOWNLOADS_PATH
        if alt.exists():
            return _read_trades_sheet(
                alt,
                "Could not find header row starting with 'Counterparty' "
                "in the FIL 'Trades' sheet",
            )
        raise


def aggregate_positions(trades: list[dict]) -> list[dict]:
    """Aggregate raw trades into net positions grouped by
    (Option Type, Strike, Option Expiry Date).

    Returns a list of position dicts with net quantities, average prices,
    and summed notionals.
    """
    positions: dict[str, dict] = {}

    for t in trades:
        # Canonicalise option type — raw data mixes "puts"/"Put"/"calls"/"Call".
        # Must normalise BEFORE building the group key, otherwise spelling
        # variants of the same option split into separate positions.
        opt_type = "Call" if "call" in str(t.get("Option Type") or "").lower() else "Put"
        strike = _safe_float(t.get("Strike"))
        expiry = str(t.get("Option Expiry Date") or "").strip()
        key = f"{opt_type}|{strike}|{expiry}"

        side_raw = str(t.get("Buy / Sell / Unwind") or "").strip().lower()
        qty = _safe_float(t.get("ETH Options"))
        premium_per = _safe_float(t.get("Premium per Contract"))
        premium_usd = _safe_float(t.get("Premium USD"))
        notional_mm = _safe_float(t.get("$ Notional (mm)"))
        ref_spot = _safe_float(t.get("Ref. Spot Price"))
        days_remaining = _safe_float(t.get("Days Remaining to Expiry"))
        pct_otm = _safe_float(t.get("% OTM"))

        if side_raw in ("buy", "long"):
            sign = 1.0
        elif side_raw in ("sell", "short"):
            sign = -1.0
        elif side_raw == "unwind":
            sign = -1.0  # unwind reduces the position
        else:
            sign = 1.0

        if key not in positions:
            positions[key] = {
                "option_type": opt_type,
                "strike": strike,
                "expiry": expiry,
                "days_remaining": days_remaining,
                "pct_otm": pct_otm,
                "ref_spot": ref_spot,
                "net_qty": 0.0,
                "total_premium_usd": 0.0,
                "total_notional_mm": 0.0,
                "trade_count": 0,
                "counterparties": set(),
            }

        pos = positions[key]
        pos["net_qty"] += sign * qty
        pos["total_premium_usd"] += sign * premium_usd
        pos["total_notional_mm"] += sign * notional_mm
        pos["trade_count"] += 1
        pos["days_remaining"] = max(pos["days_remaining"], days_remaining)
        cp = str(t.get("Counterparty") or "").strip()
        if cp:
            pos["counterparties"].add(cp)

    # Compute derived fields and serialise sets
    result: list[dict] = []
    for pos in positions.values():
        net = pos["net_qty"]
        pos["avg_premium_per_contract"] = (
            pos["total_premium_usd"] / net if net != 0 else 0.0
        )
        pos["counterparties"] = sorted(pos["counterparties"])
        pos["side"] = "Long" if net > 0 else ("Short" if net < 0 else "Flat")
        result.append(pos)

    # Sort by expiry then strike
    result.sort(key=lambda p: (p["expiry"], p["strike"]))
    return result


def read_calendar_rolls(file_path: Path | None = None) -> dict:
    """Read calendar rolls from 'Calendar rolls.xlsx'.

    Auto-detects sheet names and header rows (first row with data in each sheet).
    Returns a dict: { sheet_name: [row_dict, ...], ... }
    """
    fp = file_path or (
        Path(__file__).resolve().parents[3]
        / "data"
        / "Calendar rolls.xlsx"
    )

    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    result: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # Find the first non-empty row as the header
        headers: list[str] | None = None
        for row in rows_iter:
            if row and any(cell is not None for cell in row):
                headers = [
                    str(h).strip() if h is not None else f"col_{i}"
                    for i, h in enumerate(row)
                ]
                break

        if headers is None:
            result[sheet_name] = []
            continue

        records: list[dict] = []
        for row in rows_iter:
            if all(cell is None for cell in row):
                continue
            record = dict(zip(headers, row))
            # Normalise dates → ISO strings
            for k, v in record.items():
                if isinstance(v, datetime):
                    record[k] = v.isoformat()
                elif isinstance(v, date):
                    record[k] = v.isoformat()
            records.append(record)

        result[sheet_name] = records

    wb.close()
    return result