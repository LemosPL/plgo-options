from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl

from plgo_options.optimization.optimizer_utils import _safe_int, _safe_float


@dataclass
class Position:
    id: int
    instrument: str
    opt: str
    counterparty: str | None
    side: str
    strike: float
    expiry: str
    days_remaining: int
    net_qty: float
    iv_pct: float
    delta: float | None
    gamma: float | None
    theta: float | None
    vega: float | None
    mark_price_usd: float
    current_mtm: float
    payoff_by_horizon: dict[str, list[float]]
    mtm_by_horizon: list[float]

    @property
    def expiry_date(self) -> date:
        return date.fromisoformat(self.expiry.split("T")[0])


@dataclass
class Candidate:
    """A tradeable instrument from the vol surface."""
    expiry_code: str
    expiry_date: str
    dte: int
    strike: float
    opt: str  # "C" or "P"
    counterparty: str
    iv_pct: float
    delta: float
    gamma: float
    theta: float
    vega: float
    bs_price_usd: float

@dataclass(frozen=True)
class SpreadCandidate:
    kind: str  # "CALL_SPREAD" or "PUT_SPREAD"
    long_leg: Candidate
    short_leg: Candidate

    @property
    def expiry_code(self) -> str:
        return self.long_leg.expiry_code

    @property
    def expiry_date(self):
        return self.long_leg.expiry_date

    @property
    def dte(self) -> int:
        return self.long_leg.dte

    @property
    def counterparty(self) -> str:
        return self.long_leg.counterparty

    @property
    def opt(self) -> str:
        return self.long_leg.opt

    @property
    def strike(self) -> float:
        return self.long_leg.strike

    @property
    def width(self) -> float:
        return abs(float(self.short_leg.strike) - float(self.long_leg.strike))

    @property
    def bs_price_usd(self) -> float:
        return float(self.long_leg.bs_price_usd or 0.0) - float(self.short_leg.bs_price_usd or 0.0)

    @property
    def vega(self) -> float:
        return float(self.long_leg.vega or 0.0) - float(self.short_leg.vega or 0.0)

    @property
    def delta(self) -> float:
        return float(self.long_leg.delta or 0.0) - float(self.short_leg.delta or 0.0)

    @property
    def gamma(self) -> float:
        return float(self.long_leg.gamma or 0.0) - float(self.short_leg.gamma or 0.0)

    @property
    def iv_pct(self) -> float:
        return 0.5 * (
            float(self.long_leg.iv_pct or 0.0)
            + float(self.short_leg.iv_pct or 0.0)
        )


@dataclass(frozen=True)
class StraddleCandidate:
    kind: str  # "STRADDLE"
    call_leg: Candidate
    put_leg: Candidate

    @property
    def expiry_code(self) -> str:
        return self.call_leg.expiry_code

    @property
    def expiry_date(self):
        return self.call_leg.expiry_date

    @property
    def dte(self) -> int:
        return self.call_leg.dte

    @property
    def counterparty(self) -> str:
        return self.call_leg.counterparty

    @property
    def opt(self) -> str:
        return "STRADDLE"

    @property
    def strike(self) -> float:
        return self.call_leg.strike

    @property
    def bs_price_usd(self) -> float:
        return float(self.call_leg.bs_price_usd or 0.0) + float(self.put_leg.bs_price_usd or 0.0)

    @property
    def vega(self) -> float:
        return float(self.call_leg.vega or 0.0) + float(self.put_leg.vega or 0.0)

    @property
    def delta(self) -> float:
        return float(self.call_leg.delta or 0.0) + float(self.put_leg.delta or 0.0)

    @property
    def gamma(self) -> float:
        return float(self.call_leg.gamma or 0.0) + float(self.put_leg.gamma or 0.0)

    @property
    def theta(self) -> float:
        return float(self.call_leg.theta or 0.0) + float(self.put_leg.theta or 0.0)

    @property
    def iv_pct(self) -> float:
        return 0.5 * (
            float(self.call_leg.iv_pct or 0.0)
            + float(self.put_leg.iv_pct or 0.0)
        )


@dataclass(frozen=True)
class IronCondorCandidate:
    kind: str  # "IRON_CONDOR"
    put_low_leg: Candidate    # long put wing
    put_high_leg: Candidate   # short put body
    call_low_leg: Candidate   # short call body
    call_high_leg: Candidate  # long call wing

    @property
    def expiry_code(self) -> str:
        return self.put_low_leg.expiry_code

    @property
    def expiry_date(self):
        return self.put_low_leg.expiry_date

    @property
    def dte(self) -> int:
        return self.put_low_leg.dte

    @property
    def counterparty(self) -> str:
        return self.put_low_leg.counterparty

    @property
    def opt(self) -> str:
        return "IRON_CONDOR"

    @property
    def strike(self) -> float:
        return 0.5 * (
            float(self.put_high_leg.strike or 0.0)
            + float(self.call_low_leg.strike or 0.0)
        )

    @property
    def bs_price_usd(self) -> float:
        return (
            float(self.put_low_leg.bs_price_usd or 0.0)
            - float(self.put_high_leg.bs_price_usd or 0.0)
            - float(self.call_low_leg.bs_price_usd or 0.0)
            + float(self.call_high_leg.bs_price_usd or 0.0)
        )

    @property
    def vega(self) -> float:
        return (
            float(self.put_low_leg.vega or 0.0)
            - float(self.put_high_leg.vega or 0.0)
            - float(self.call_low_leg.vega or 0.0)
            + float(self.call_high_leg.vega or 0.0)
        )

    @property
    def delta(self) -> float:
        return (
            float(self.put_low_leg.delta or 0.0)
            - float(self.put_high_leg.delta or 0.0)
            - float(self.call_low_leg.delta or 0.0)
            + float(self.call_high_leg.delta or 0.0)
        )

    @property
    def gamma(self) -> float:
        return (
            float(self.put_low_leg.gamma or 0.0)
            - float(self.put_high_leg.gamma or 0.0)
            - float(self.call_low_leg.gamma or 0.0)
            + float(self.call_high_leg.gamma or 0.0)
        )

    @property
    def theta(self) -> float:
        return (
            float(self.put_low_leg.theta or 0.0)
            - float(self.put_high_leg.theta or 0.0)
            - float(self.call_low_leg.theta or 0.0)
            + float(self.call_high_leg.theta or 0.0)
        )

    @property
    def iv_pct(self) -> float:
        return 0.25 * (
            float(self.put_low_leg.iv_pct or 0.0)
            + float(self.put_high_leg.iv_pct or 0.0)
            + float(self.call_low_leg.iv_pct or 0.0)
            + float(self.call_high_leg.iv_pct or 0.0)
        )


POSITIONS_DIR = Path(__file__).resolve().parent.parent.parent.parent / "data" / "positions"
def load_positions_from_latest_xlsx(positions_dir: Path = POSITIONS_DIR) -> list[Position]:
    """Load positions from the most recent .xlsx file in data/positions."""
    if not positions_dir.exists() or not positions_dir.is_dir():
        return []

    xlsx_files = sorted(
        positions_dir.glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not xlsx_files:
        return []

    latest_file = xlsx_files[0]
    wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

    # Use the first sheet unless you want to pin this to a specific name.
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    headers: list[str] | None = None
    positions: list[Position] = []

    for row in rows:
        if row and any(cell is not None for cell in row):
            headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(row)]
            break

    if headers is None:
        wb.close()
        return []

    for row in rows:
        if not row or all(cell is None for cell in row):
            continue

        record = dict(zip(headers, row))

        instrument = str(record.get("Instrument") or "").strip()
        if not instrument:
            continue

        opt = str(record.get("Type") or "").strip().upper()
        if opt in {"CALL", "C"}:
            opt = "C"
        elif opt in {"PUT", "P"}:
            opt = "P"

        side = str(record.get("Side") or "").strip()
        counterparty = str(record.get("Counterparty") or "brokerage").strip() or "brokerage"
        expiry = str(record.get("Expiry") or "").strip()

        # Build a stable synthetic ID if needed
        raw_id = record.get("ID")
        pos_id = _safe_int(raw_id, default=len(positions) + 1)

        positions.append(
            Position(
                id=pos_id,
                instrument=instrument,
                opt=opt,
                counterparty=counterparty,
                side=side,
                strike=_safe_float(record.get("Strike")),
                expiry=expiry,
                days_remaining=_safe_int(record.get("DTE")),
                net_qty=_safe_float(record.get("Qty")),
                iv_pct=_safe_float(record.get("IV%")),
                delta=record.get("Delta"),
                gamma=record.get("Gamma"),
                theta=record.get("Theta"),
                vega=record.get("Vega"),
                mark_price_usd=_safe_float(record.get("Mark Price")),
                current_mtm=_safe_float(record.get("MTM")),
                payoff_by_horizon={},
                mtm_by_horizon=[],
            )
        )

    wb.close()
    return positions